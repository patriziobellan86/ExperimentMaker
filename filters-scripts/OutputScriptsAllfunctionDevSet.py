import codecs
import tkinter as tk
from tkinter import filedialog
from pathlib import Path
from copy import deepcopy
from itertools import chain
from openpyxl import Workbook
from openpyxl import load_workbook
import os

_ACTIVITY = 'Activity'
_PARTICIPANT = "Participant"
_ACTIVITYPARTICIPANT = "Activity-Participant"
_FLOW = "Flow"

_MAX_EMPTY_ROWS = 5  # number of empty row between items

#  SHEET' SETTINGS
_LABEL_COLUMN = 'B'
_ISTANCE_COLUMN = 'C'



def OrderResults(results):
    results = deepcopy(results)
    results_keys = list(results.keys())
    keys_step_n_pairs = [(k, results[k]['step']) for k in results_keys]
    keys_step_n_pairs_ordered = sorted(keys_step_n_pairs, key=lambda x: x[1])
    keys_ordered = [k for k, s in keys_step_n_pairs_ordered]
    results_ordered = {k: results[k] for k in keys_ordered}
    return results_ordered

def GetResultItem(step_name,
                  results):
    results = deepcopy(results)

    for key in results:
        if key == step_name:
            return results[key]

def GetResultsContent(result_step):
    #  GET THE CONTENT FROM A RESULTS_STEP

    results_content = list()
    if result_step['results-filtered']:
        for item in result_step['results-filtered'][-1]:
            results_content.append(item['result'])
        #  flatten the list
        results_content = list(chain(*results_content))

    else:
        for item in result_step['results-raw']:
            results_content.append(item['result'])
        # if len(results) > 1:
        #     results = list(chain(*results))
    return results_content

def GetResultOutputContent(step_name,
              results):
    #  GET THE OUTPUT FROM A LIST OF RESULTS
    # results = deepcopy(results)
    # return the results raw if no filter is applied.
    results_content = list()
    if results[step_name]['results-filtered']:
        for item in results[step_name]['results-filtered'][-1]:
            results_content.append(item['result'])
        #  flatten the list
        results_content = list(chain(*results_content))

    else:
        for item in results[step_name]['results-raw']:
            results_content.append(item['result'])
        # if len(results) > 1:
        #     results = list(chain(*results))
    return results_content


class output_base:
    def __init__(self):
        self.filaname = None

    def ask_save_filename(self):
        app = tk.Tk()
        app.geometry('0x0')
        filename = Path(filedialog.asksaveasfilename())
        if filename.name:
            self.filename = str(filename.absolute())
            return self.filename
        else:
            return False

    def Parse(self):
        raise NotImplementedError

class CreateOntologyIstance(output_base):
    def __init__(self,
                 step_name,
                 ontology_params):
        super(CreateOntologyIstance, self).__init__()

        self.step_name = step_name
        self.ontology_params = ontology_params

    def Parse(self,
              results):
        results = OrderResults(results)
        step_results = GetResultItem(self.step_name, results)
        results_content = GetResultsContent(step_results)
        return results_content

class ParseActivityIstances(CreateOntologyIstance):
    def __init__(self):
        super(ParseActivityIstances, self).__init__('step1', {})
    #
class ParseActivityParticipantIstances(CreateOntologyIstance):
    def __init__(self):
        super(ParseActivityParticipantIstances, self).__init__('step1', {})
    #
    # def Parse(self,
    #           results):
    #     # if not self.ask_save_filename(): return
    #     super_step_results = super(ParseActivityIstances, self).Parse(results)
    #     activities_list = GetResultsContent(super_step_results)
    #     print('--- activity list ---')
    #     print(activities_list)
    #     # codecs.open(self.filename, 'w', 'utf8').write('\n\n--- activity list ---\n')
    #     # codecs.open(self.filename, 'a', 'utf8').writelines(activities_list)
    #     return activities_list

class ParseParticipantIstances(CreateOntologyIstance):
    def __init__(self):
        super(ParseParticipantIstances, self).__init__('step2', {})

    # def Parse(self,
    #           results):
    #     # if not self.ask_save_filename(): return
    #     super_ste_results = super(ParseParticipantIstances, self).Parse(results)
    #     participant_list = GetResultsContent(super_ste_results)
    #     print('--- participant list ---')
    #     print(participant_list)
    #     # codecs.open(self.filename, 'w', 'utf8').write('\n\n--- participant list ---\n')
    #     # codecs.open(self.filename, 'a', 'utf8').writelines(participant_list)
    #     print('--- unique participants ---')
    #     unique_participant = list(set(participant_list))
    #     print(unique_participant)
    #     # codecs.open(self.filename, 'a', 'utf8').write('\n\n--- unique participants ---\n')
    #     # codecs.open(self.filename, 'a', 'utf8').writelines(participant_list)
    #     return participant_list


class ParseFlowIstances(CreateOntologyIstance):
    def __init__(self):
        super(ParseFlowIstances, self).__init__('step1', {})
    #
class ParseInstance(output_base):

    def __init__(self):
        super(ParseInstance, self).__init__()

    def Parse(self,
              results):
        self.activity = ParseActivityIstances().Parse(results)
        self.participant = ParseParticipantIstances().Parse(results)
        self.activityparticipant = ParseActivityParticipantIstances().Parse(results)
        self.flow = ParseFlowIstances().Parse(results)



class ExportInExcelFile(output_base):
    ACTIVITY = 'Activity'
    PARTICIPANT = "Participant"
    ACTIVITYPARTICIPANT = "Activity-Participant"
    FLOW = "Flow"

    MAX_EMPTY_ROWS = 5 #  number of empty row between items

    #  SHEET' SETTINGS
    _LABEL_COLUMN = 'B'
    _ISTANCE_COLUMN = 'C'

    def __init__(self):
        super(ExportInExcelFile, self).__init__()
        self.wb = None
        self.row_index = {self.ACTIVITY: 1,
                          self.PARTICIPANT: 1,
                          self.ACTIVITYPARTICIPANT: 1,
                          self.FLOW: 1}

    def LoadExcelFile(self):
        self.wb = load_workbook(self.filename)
        #  read indexes
        self.row_index = {self.ACTIVITY: self._get_row_index_activity(),
                          self.PARTICIPANT: self._get_row_index_participant(),
                          self.ACTIVITYPARTICIPANT: self._get_row_index_activity_participant(),
                          self.FLOW: self._get_row_index_flow(),
                          }

    def _get_index_empty(self, sheet_name):
        ws = self.wb[sheet_name]
        index = 1  #  index STARTS from 1.

        max_empty_row_counter = 0
        while True:
            row = ws[index]
            if not any([cell.value for cell in row]):
                max_empty_row_counter += 1
                if max_empty_row_counter == self.MAX_EMPTY_ROWS:
                    return index
            else:
                print(row)
                max_empty_row_counter = 0
            index += 1

    def _get_row_index_activity(self):
        return self._get_index_empty(self.ACTIVITY)

    def _get_row_index_participant(self):
        return self._get_index_empty(self.PARTICIPANT)

    def _get_row_index_activity_participant(self):
        return self._get_index_empty(self.ACTIVITYPARTICIPANT)

    def _get_row_index_flow(self):
        return self._get_index_empty(self.FLOW)

    def write_activity(self,
                        activity_list):
        self._write_istance_list(self.ACTIVITY,
                                 activity_list)
    def write_participant(self,
                        participant_list):
        self._write_istance_list(self.PARTICIPANT,
                                 participant_list)
    def write_activityparticipant(self,
                        activityparticipant_list):
        self._write_istance_list(self.ACTIVITYPARTICIPANT,
                                 activityparticipant_list)
    def write_flow(self,
                        flow_matrix):

        print('change this! flow is a matrix!')
        self._write_istance_list(self.FLOW,
                                 flow_matrix)

    def _write_istance_list(self,
                            label,
                            istances):
        row_index = self.row_index[label]
        ws = self.wb[label]

        key = '{col}{row}'.format(col=self._LABEL_COLUMN,
                                  row=row_index)
        ws[key] = Path(self.filename).name
        row_index += 1

        for instance in istances:
            key = '{col}{row}'.format(col=self._ISTANCE_COLUMN,
                                      row=row_index)
            ws[key] = instance
            row_index += 1

        self.row_index[label] = row_index

    def CreateExcelFile(self):
        self.wb = Workbook()
        self.wb.create_sheet(self.ACTIVITY)
        self.wb.create_sheet(self.PARTICIPANT)
        self.wb.create_sheet(self.ACTIVITYPARTICIPANT)
        self.wb.create_sheet(self.FLOW)

    def SaveExcelFile(self):
        self.wb.save(self.filename)

    def Parse(self,
              results):
        if self.ask_save_filename():
            print('selected filename: ', self.filename)
        if os.path.exists(self.filename):
            self.LoadExcelFile()
        else:
            self.CreateExcelFile()

        #  add Activity
        activity = ParseActivityIstances().Parse(results)
        self.write_activity(activity)
        #  add Participant
        participant = ParseParticipantIstances().Parse(results)
        self.write_participant(participant)


        self.SaveExcelFile()
#
# if __name__ == '__main__':
#     import json
#     from promptdesignerdataset.dataset import LoadJsonData
#     results_file = 'results.json'
#     results = LoadJsonData(results_file)
#     filter = ParseActivityIstances().Parse
#     res= filter(results)
#     print(res)
#
#     filter = ParseParticipantIstances().Parse
#     res= filter(results)
#     print(res)
#
#
#     filter = ExportInExcelFile().Parse
#     res= filter(results)
#     print(res)
#
#     filter = ParseInstance().Parse
#     res= filter(results)
#     print(res)
